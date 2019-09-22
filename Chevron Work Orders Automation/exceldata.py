import openpyxl
import datetime
import math
import smtplib
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart

# sms initial variables
email = 'pythonsmstest12@gmail.com'
pas = 'hackrice123'
sms_gateway = '5108945257@txt.att.net'
smtp = 'smtp.gmail.com'
port = 587
server = smtplib.SMTP(smtp,port)

# inputs the workbook and all of the sheets
wb = openpyxl.load_workbook('workorder.xlsx', read_only=True, data_only=True)
wb2 = openpyxl.load_workbook('testdata1.xlsx', read_only=True, data_only=True)
equipment_details = wb['Equipment Details']
worker_details = wb['Worker Details']
facility_details = wb['Facility Details']
work_order_examples = wb2['Sheet2']

# worker list
workers = []
for row in range(2, work_order_examples.max_row + 1):

    # exit out of the for loop if the row is empty
    if worker_details.cell(row=row, column=2).value is None:
        break

    # set shifts. Morning = True, Evening = False
    if str(worker_details.cell(row=row, column=4).value) == 'Morning':
        shift = True
    else:
        shift = False

    # assign specialization groups

    '''
    Sensor, Security, Networking : spec = 1
    Pump, HVAC : spec = 2
    Vehicle : spec = 3
    Conveyor, Seperator : spec = 4
    '''

    if str(worker_details.cell(row=row, column=3).value) == 'Sensor, Security, Networking':
        spec = 1
    elif str(worker_details.cell(row=row, column=3).value) == 'Pump, HVAC':
        spec = 2
    elif str(worker_details.cell(row=row, column=3).value) == 'Vehicle':
        spec = 3
    elif str(worker_details.cell(row=row, column=3).value) == 'Conveyor, Seperator':
        spec = 4

    # fill in the dictionary
    worker_detail = {
        str(worker_details.cell(row=row, column=2).value): [spec, shift]}
    workers.append(worker_detail)

# work orders
work_order = []
for row in range(3, work_order_examples.max_row + 1):

    # exit out of the for loop if the row is empty
    if work_order_examples.cell(row=row, column=2).value is None:
        break

    # facility number
    if str(work_order_examples.cell(row=row, column=3).value) == 'Fac1':
        facnum = 1
    elif str(work_order_examples.cell(row=row, column=3).value) == 'Fac2':
        facnum = 2
    elif str(work_order_examples.cell(row=row, column=3).value) == 'Fac3':
        facnum = 3
    elif str(work_order_examples.cell(row=row, column=3).value) == 'Fac4':
        facnum = 4
    elif str(work_order_examples.cell(row=row, column=3).value) == 'Fac5':
        facnum = 5

    # calculate the minutes from the current time from the timestamp
    if '/' in str(work_order_examples.cell(row=row, column=8).value):
        time_datetime = datetime.datetime.strptime(str(work_order_examples.cell(row=row, column=8).value),
                                                   '%m/%d/%Y %H:%M:%S')
    elif '-' in str(work_order_examples.cell(row=row, column=8).value):
        time_datetime = datetime.datetime.strptime(str(work_order_examples.cell(row=row, column=8).value),
                                                   '%Y-%m-%d %H:%M:%S')
    timediff = datetime.datetime.now() - time_datetime
    mins = (timediff.days * 24 * 60) + round((timediff.seconds / 60))
    dummy = 1


    # fill in the dictionary
    working_dict = {'fac_num': facnum,
                    'eq_type': str(work_order_examples.cell(row=row, column=4).value),
                    'priority': int(work_order_examples.cell(row=row, column=6).value),
                    'time_to_complete': int(work_order_examples.cell(row=row, column=7).value),
                    'mins_from_now': mins}
    work_order.append(working_dict)

work_order_types = [[], [], [], [], []]

# looked up in google maps for the distance
# between the two places, and added 5 minutes
# for parking

DRIVE_TIMES = [[0, 18, 18, 17, 21], [19, 0, 8, 10, 14], [19, 8, 0, 10, 11], [18, 10, 9, 0, 10], [18, 9, 8, 11, 0]]

TYPE_0 = ["Sensor", "Security", "Networking"]
TYPE_1 = ["Pump", "HVAC"]
TYPE_2 = ["Vehicle"]
TYPE_3 = ["Conveyer", "Seperator"]
TYPE_4 = ["Compressor", "Electricity"]
TYPES = [TYPE_0, TYPE_1, TYPE_2, TYPE_3, TYPE_4]


def filter_work_order(task):
    for index in range(len(TYPES)):
        if task['eq_type'] in TYPES[index]:
            work_order_types[index].append(task)
    return None


def filter_all_work_orders(tasklist):
    for task in tasklist:
        filter_work_order(task)
    return None

filter_all_work_orders(work_order)

def drive_time(location1, location2):
    return DRIVE_TIMES[location1][location2]


def time_score(work_orders, ordernum):
    task_time = work_orders[ordernum]['time_to_complete'] * 60
    from_facility = work_orders[ordernum - 1]['fac_num'] - 1
    to_facility = work_orders[ordernum]['fac_num'] - 1
    travel_time = drive_time(from_facility, to_facility)
    return task_time + travel_time


def total_time(work_orders):
    sum1 = 0
    for index in range(len(work_orders)): # order in work_orders:
        sum1 += time_score(work_orders, index)
    return sum1


def priority_score(item):
    priority = item["priority"]
    time_passed = item['mins_from_now']
    modifier = math.floor(time_passed/1440)
    if modifier == 0:
        modifier = 0.000000001
    return float(priority / modifier)


def permutation(lst):
    if len(lst) == 0:
        return []
    if len(lst) == 1:
        return [lst]
    l = []
    for i in range(len(lst)):
        m = lst[i]
        remLst = lst[:i] + lst[i + 1:]
        for p in permutation(remLst):
            l.append([m] + p)
    return l


def find_min_time(work_orders):
    perms = permutation(work_orders)
    min = total_time(perms[0])
    for perm in perms:
        if total_time(perm) < min:
            min = total_time(perm)
    return min


def optimizetime(work_orders):
    min1 = find_min_time(work_orders)
    goodperms = []
    for perm in permutation(work_orders):
        if total_time(perm) < 1.05 * min1:
            goodperms.append(perm)
    return goodperms


def find_min_priority(permlist):
    bestperm = permlist[0]
    min = priority_score(bestperm)
    for perm in permlist:
        if priority_score(perm) < min:
            min = priority_score(perm)
            bestperm = perm
    return min


def bestpriorityorder(work_orders):
    work_orders.sort(key=priority_score)
    return work_orders


def rsquared(perm, work_orders):
    sum1 = 0
    for index in range(len(perm)):
        sum1 += ((priority_score(perm[index]) - priority_score(bestpriorityorder(work_orders)[index])) ** 2)
    return sum1


def minimize_r2(perms, work_orders):
    bestperm = perms[0]
    for perm in perms:
        if rsquared(perm, work_orders) < rsquared(bestperm, work_orders):
            bestperm = perm
    return bestperm


def the_best_function_ever(work_orders):
    contracted_work_orders = optimizetime(work_orders)
    the_best_order_ever = minimize_r2(contracted_work_orders, work_orders)
    return the_best_order_ever

order = []
for i in range(5):
    order.append(the_best_function_ever(work_order_types[i]))
test = 1

#
# sms
#

server.starttls()
server.login(email, pas)

msg = MIMEMultipart()
msg['From'] = 'Chevron'
msg['To'] = sms_gateway

msg['Subject'] = 'Work Assignment'
tasklist = []
for i in range(len(order[1])):
    task = ('Task #' + str(i + 1) + ':\nFacility: ' + str(order[0][i]['fac_num']) + '\nEquipment: '
            + order[0][i]['eq_type'] + '\nPriority: ' + str(order[0][i]['priority']) + '\n')
    tasklist.append(task)
body = '\n'.join(tasklist)
body = '\n' + body

msg.attach(MIMEText(body, 'plain'))

sms = msg.as_string()

server.sendmail(email, sms_gateway,sms)

server.quit()
dummy = 1